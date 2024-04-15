'''
2. Посчитайте суммарную стоимость оборудования (лист "Реестр оборудования") по Магазинам в зависимоти от года выпуска (зелёные ячейки) и общие суммы (жёлтые ячейки).


Через Питон не получилось выполнить это задание, так как некоторые поля с ценами ссылались на предыдущее поля с его ценой(например, такого формата =H955+7854), поэтому выполнила задание с помощью самого Excel и полученные ответы записала в файле
'''



import openpyxl


wb = openpyxl.load_workbook('excel.xlsx')
ws = wb['Математические'] 

shop_names = []

# Перебираем каждую строку с 21 по 35 строку включительно и сохраняем значение из ячейки B в список
for row in range(21, 36):
    shop_name = ws.cell(row=row, column=3).value
    shop_names.append(shop_name)

# print(shop_names)

ws_2 = wb['Реестр оборудования']


sum_cost_equipment = {}

# Перебираем каждую строку в файле, начиная со 2 строки и вытягиваем только значения данных
for row in ws_2.iter_rows(values_only=True):
    shop = row[0]
    year = row[5]
    price = row[7]


    # Создаем проверку на тип данных года. Нужно только числовые данные вытянуть.
    if not isinstance(year, int):
        continue

    if not isinstance(price, (float, int)):
       continue
        
    
    # print(price)

    # Если оборудование было выпущено до 2015 года, увеличиваем счетчик для этого оборудования
    if shop in shop_names:
        if year <= 2015:
            if shop not in sum_cost_equipment:
                sum_cost_equipment[shop] = 0
            sum_cost_equipment[shop] += price

print(sum_cost_equipment)