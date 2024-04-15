'''
1. Посчитайте кол-во оборудования (лист "Реестр оборудования") по годам выпуска (зелёные ячейки) и суммарно в каждом году (жёлтые ячейки). 
'''

import openpyxl

wb = openpyxl.load_workbook('excel.xlsx')
ws = wb['Реестр оборудования']

equipment_by_year = {}

# Создаем переменные для общего количества оборудования, выпущенного до 2015 года
equipment_until_2015 = {}
total_count_2015 = 0

# Перебираем каждую строку в файле, начиная со 2 строки и вытягиваем только значения данных
for row in ws.iter_rows(min_row=2, values_only=True):
    year = row[5]
    equipment_name = row[1]
    count = 1

    if not isinstance(year, int):
        continue

    # Создаем отдельное условие, чтобы записать в одну колонку все оборудования, выпущенные за 2015
    if year <= 2015:
        equipment_until_2015[equipment_name] = equipment_until_2015.get(equipment_name, 0) + count
        total_count_2015 += count
        equipment_by_year['<=2015'] = equipment_until_2015
        continue

    # Если год еще не встречался, создаем ключ с этим годом
    if year not in equipment_by_year:
        equipment_by_year[year] = {}
        total_equipment = 0

    # Увеличиваем счетчик для этого оборудования и года
    equipment_by_year[year][equipment_name] = equipment_by_year[year].get(equipment_name, 0) + count
    equipment_by_year[year]['Общее количество'] = equipment_by_year[year].get('Общее количество', 0) + count


equipment_by_year['<=2015']['Общее количество'] = total_count_2015


# Создаем файл excel, куда будем загружать конечный результат. Это позволит не вносить вручную данные и не ошибиться во время переноса значений!

wb_out = openpyxl.Workbook()
ws_out = wb_out.active

# Добавляем заголовки
ws_out.cell(row=1, column=1, value='Название оборудования')

# Записываем названия оборудования по вертикали
row_index = 2
for equipment_name in equipment_until_2015:
    ws_out.cell(row=row_index, column=1, value=equipment_name)
    row_index += 1

# Записываем данные по горизонтали
column_index = 2
for year, equipment_data in equipment_by_year.items():
    # Добавляем заголовок для года
    ws_out.cell(row=1, column=column_index, value=year)
    
    # Записываем количество оборудования по каждому виду
    row_index = 2
    for equipment_name in equipment_until_2015:
        count = equipment_data.get(equipment_name, 0)
        ws_out.cell(row=row_index, column=column_index, value=count)
        row_index += 1

    column_index += 1
# Сохраняем файл
wb_out.save('output.xlsx')

print("Данные успешно записаны в файл output.xlsx")




# Можно также в формате CSV данные записать. 
import csv
with open('output.csv', 'w', newline='', encoding='utf-8') as csvfile:

    csvwriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # Записываем заголовок
    csvwriter.writerow(['Год выпуска', 'Название оборудования', 'Количество'])

    # Записываем результаты
    for year, equipment_data in equipment_by_year.items():
        for equipment_name, count in equipment_data.items():
            if equipment_name != 'Общее количество':
                csvwriter.writerow([year, equipment_name, count])
        if 'Общее количество' in equipment_data:
            csvwriter.writerow([year, 'Общее количество оборудования', equipment_data['Общее количество']])

print("Данные успешно записаны в файл output.csv")
