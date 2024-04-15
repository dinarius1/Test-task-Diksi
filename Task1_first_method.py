'''
Первый способ решения задачи с использоанием библиотеки Selenium. Скрпит решает задание за 2090 милисекунд. 
'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Загружаем наш файл с данными о пользователях
wb = openpyxl.load_workbook('challenge.xlsx')
ws = wb.active

# Получаем названия столбцов из первой строки. 
col_names = [ws.cell(row=1, column=j).value.strip() for j in range(1, 8)]

# print(col_names)

# exit()

# Запускаем браузер
driver = webdriver.Chrome()
driver.get("https://www.rpachallenge.com/")

# Нажимаем кнопку "start"
start_button = driver.find_element(By.CLASS_NAME, "uiColorButton")
start_button.click()

# Создаем словарь, чтобы сопоставить названия полей из excel с соответствующими элементами формы на сайте
field_mapping = {
    "First Name": "//input[@ng-reflect-name='labelFirstName']",
    "Last Name": "//input[@ng-reflect-name='labelLastName']",
    "Company Name": "//input[@ng-reflect-name='labelCompanyName']",
    "Role in Company": "//input[@ng-reflect-name='labelRole']",
    "Address": "//input[@ng-reflect-name='labelAddress']",
    "Email": "//input[@ng-reflect-name='labelEmail']",
    "Phone Number": "//input[@ng-reflect-name='labelPhone']"
}

# Проходим по строкам файла excel и заполняем форму
for i in range(2, 12):
    vals = {col_names[j-1]: ws.cell(row=i, column=j).value for j in range(1, 8)}

    # Ожидаем элементы формы
    wait = WebDriverWait(driver, 10)

    # Заполняем форму данными из excel
    for field_name, field_value in vals.items():
        input_element = wait.until(EC.presence_of_element_located((By.XPATH, field_mapping[field_name])))
        input_element.send_keys(str(field_value))

    # Нажимаем кнопку "submit"
    submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='submit']")))
    submit_button.click()

# Ждем, чтобы посмотреть результат
time.sleep(15)

driver.quit()



